#!/usr/bin/env python3
"""
fetch_cfe_tarifa_ts.py
======================
Extracts CFE Suministro Básico time-series data from official CRE/CNE workbooks.

PART 1 — USERS (from "Memoria de cálculo de las Tarifas de Operación del SSB")
PART 2 — ENERGY (from "Memoria de cálculo de las Tarifas Finales del Suministro Básico")

SOURCE (open-data attribution):
  https://www.gob.mx/cne/articulos/memorias-de-calculo-de-las-tarifas-electricas

Energy files download URL pattern (CNE, TLS chain incomplete → use verify=False / curl -k):
  https://www.cne.gob.mx/da/Memorias%20de%20c%C3%A1lculo%20de%20tarifas%20finales%20del%20Suministro%20B%C3%A1sico/TarifasFinalesdeSuministroBasico{YYYY}.xlsx

Usage:
    python3 fetch_cfe_tarifa_ts.py

Outputs:
    ../public/data/cfe_users_ts.json
    ../public/data/cfe_energy_ts.json
"""

import os
import json
import urllib.request
import ssl
from pathlib import Path
import openpyxl

# ─── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR.parent / "public" / "data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Adjust these paths if your local files live elsewhere
DOWNLOADS = Path(os.environ.get("CFE_DOWNLOADS", Path.home() / "Downloads"))
FINALES_CACHE = DOWNLOADS / "cfe_finales"
FINALES_CACHE.mkdir(exist_ok=True)

SOURCE_URL = "https://www.gob.mx/cne/articulos/memorias-de-calculo-de-las-tarifas-electricas"
FINALES_BASE = (
    "https://www.cne.gob.mx/da/"
    "Memorias%20de%20c%C3%A1lculo%20de%20tarifas%20finales%20del%20Suministro%20B%C3%A1sico/"
    "TarifasFinalesdeSuministroBasico{year}.xlsx"
)

# ─── Canonical division name map ─────────────────────────────────────────────
# Maps raw strings from the workbooks → canonical map keys.
DIVISION_NORM = {
    "Baja California": "Baja California",
    "Baja California Sur": "Baja California Sur",
    "Bajío": "Bajío",
    "Centro Occidente": "Centro Occidente",
    "Centro Oriente": "Centro Oriente",
    "Centro Sur": "Centro Sur",
    "Golfo Centro": "Golfo Centro",
    "Golfo Norte": "Golfo Norte",
    "Jalisco": "Jalisco",
    "Noroeste": "Noroeste",
    "Norte": "Norte",
    "Oriente": "Oriente",
    "Peninsular": "Peninsular",
    "Sureste": "Sureste",
    "Valle de México Centro": "Valle de México Centro",
    "Valle de México Norte": "Valle de México Norte",
    "Valle de México Sur": "Valle de México Sur",
    # Aliases observed in some files
    "Valle de Mexico Centro": "Valle de México Centro",
    "Valle de Mexico Norte": "Valle de México Norte",
    "Valle de Mexico Sur": "Valle de México Sur",
    "BCS": "Baja California Sur",
    "BC": "Baja California",
}

CANONICAL_DIVISIONS = sorted(set(DIVISION_NORM.values()))

TARIFA_CODES = ["DB1", "DB2", "PDBT", "GDBT", "GDMTH", "GDMTO", "DIST", "DIT",
                "RABT", "RAMT", "APBT", "APMT"]


def norm_div(raw):
    """Normalize a raw division string to canonical form, or return None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s in DIVISION_NORM:
        return DIVISION_NORM[s]
    # Fuzzy fallback: accent-insensitive
    for k, v in DIVISION_NORM.items():
        if k.lower().replace("é", "e").replace("í", "i").replace("ó", "o") == \
                s.lower().replace("é", "e").replace("í", "i").replace("ó", "o"):
            return v
    return None


# ─── Nested dict helpers ──────────────────────────────────────────────────────

def set_nested(d, div, tarifa, year, value):
    d.setdefault(div, {}).setdefault(tarifa, {})[str(year)] = round(value, 2) if value is not None else None


# ─── PART 1: USERS ────────────────────────────────────────────────────────────

def parse_users_2017_2021(path):
    """
    Sheet '3. Proyección Usuarios':
      Row 4: header — cols 1=Aux, 3=Tarifa, 4=Concepto, 5=División, 6..N=monthly dates
      Data rows: Concepto='Usuarios', values in thousands? No—raw count.
      Coverage: 2016-01 through 2018-12 (36 months in cols 6-41).
    Returns dict: {div: {tarifa: {year: value}}} using December value for each year.
    """
    print(f"  Parsing 2017-2021 file: {path.name}")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["3. Proyección Usuarios"]

    # Read header
    header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    date_cols = {}  # col_idx (0-based) -> datetime
    for i, v in enumerate(header):
        if hasattr(v, 'year'):
            date_cols[i] = v

    # Build year -> [col_indices with that year]
    year_cols = {}
    for idx, dt in date_cols.items():
        year_cols.setdefault(dt.year, []).append(idx)

    results = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        tarifa = row[2]
        concepto = row[3]
        div_raw = row[4]
        if concepto != "Usuarios":
            continue
        if tarifa not in TARIFA_CODES:
            continue
        div = norm_div(div_raw)
        if div is None:
            print(f"    WARNING: unmapped division '{div_raw}'")
            continue
        # For each year in coverage, take December value (last month of year)
        for year, col_list in year_cols.items():
            # pick December (month==12) if available, else last
            dec_cols = [c for c in col_list if date_cols[c].month == 12]
            use_cols = dec_cols if dec_cols else col_list[-1:]
            for c in use_cols:
                val = row[c]
                if val is not None:
                    set_nested(results, div, tarifa, year, float(val))
    wb.close()
    print(f"    → {len(results)} divisions, years {sorted(set(y for d in results.values() for t in d.values() for y in t))}")
    return results


def parse_users_matrix(path, sheetname, year):
    """
    Two layout variants:

    VARIANT A — 2022, 2023 ('5. Usuarios'):
      Row 11: division names across columns (col 3..18)
      Row 12+: col 2 = tarifa code, col 3..18 = user values per division

    VARIANT B — 2024, 2025, 2026 ('2. Usuarios' / '2. Usuarias'):
      Row 10: tarifa codes across columns (col 3..14)
      Row 12+: col 2 = division name, col 3..14 = user values per tarifa

    Detection: if the row containing 'DB1' has a canonical division name elsewhere
    in the SAME row → Variant A; otherwise Variant B.
    Returns {div: {tarifa: {year: value}}}
    """
    print(f"  Parsing {year} users from '{sheetname}' in {path.name}")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheetname]

    all_rows = list(ws.iter_rows(min_row=1, max_row=40, max_col=20, values_only=True))

    # Find the row that contains 'DB1'
    db1_row_idx = None
    for i, row in enumerate(all_rows):
        if 'DB1' in row:
            db1_row_idx = i
            break
    if db1_row_idx is None:
        print(f"    ERROR: could not find DB1 in {path.name}/{sheetname}")
        wb.close()
        return {}

    db1_row = all_rows[db1_row_idx]

    # Detect variant:
    # Variant A (2022-2023): DB1 is in col 2 of a DATA row (col 1=tarifa, col 2+ = div values)
    #   The row ABOVE contains division names as column headers.
    # Variant B (2024-2026): DB1 is in col 3+ of a HEADER row (col 2=division in data rows below)
    #   The DB1 row IS the tarifa header; the next row has descriptions.
    #
    # Detection: check if DB1 is in col index 1 (Variant A) or col index >= 2 (Variant B)
    db1_col_idx = next(i for i, v in enumerate(db1_row) if v == 'DB1')
    is_variant_a = (db1_col_idx == 1)  # DB1 is the tarifa label in col 2 (0-indexed: 1)

    results = {}

    if is_variant_a:
        # Variant A: row ABOVE db1_row_idx has division names as column headers
        div_header = all_rows[db1_row_idx - 1] if db1_row_idx > 0 else []
        col_div = {}  # col_index -> canonical div name
        for i, v in enumerate(div_header):
            if i < 2:
                continue
            d = norm_div(str(v)) if v is not None else None
            if d:
                col_div[i] = d

        # Data rows start at db1_row_idx (DB1 is already a data row)
        for row in all_rows[db1_row_idx:]:
            if row[1] is None:
                continue
            tarifa = str(row[1]).strip()
            if tarifa not in TARIFA_CODES:
                continue
            for col_idx, div in col_div.items():
                val = row[col_idx]
                if val is not None and isinstance(val, (int, float)):
                    set_nested(results, div, tarifa, year, float(val))
    else:
        # Variant B: DB1 row is the TARIFA header row; col 2+ = tarifas
        # Data rows below: col 2 = division name, col 3+ = values
        tarifa_header = db1_row
        col_tarifa = {}  # col_index -> tarifa code
        for i, v in enumerate(tarifa_header):
            if v in TARIFA_CODES:
                col_tarifa[i] = v

        for row in all_rows[db1_row_idx + 1:]:
            if row[1] is None:
                continue
            div_raw = str(row[1]).strip()
            if div_raw.upper().startswith("TOTAL") or div_raw.startswith("Fuente"):
                continue
            if len(div_raw) > 60:
                continue  # description row
            div = norm_div(div_raw)
            if div is None:
                print(f"    WARNING: unmapped division '{div_raw}'")
                continue
            for col_idx, tarifa in col_tarifa.items():
                val = row[col_idx]
                if val is not None and isinstance(val, (int, float)):
                    set_nested(results, div, tarifa, year, float(val))

    wb.close()
    print(f"    → {len(results)} divisions")
    return results


def merge_users(base, override):
    """Merge override into base; override wins on same year."""
    for div, tarifas in override.items():
        for tarifa, years in tarifas.items():
            for year, val in years.items():
                set_nested(base, div, tarifa, year, val)


def build_users_ts():
    """Main users extraction. Returns merged dict."""
    data = {}

    # 2016-2018 from 2017-2021 file
    path = DOWNLOADS / "MemoriaCalculoTarifaOperacionSuministroServiciosBasicos2017_2021 (1).xlsx"
    if path.exists():
        merge_users(data, parse_users_2017_2021(path))
    else:
        print(f"  MISSING: {path}")

    # 2022-2026: annual matrix files (newer overrides older for same year)
    annual_files = [
        (DOWNLOADS / "MemoriaTarifasdeOperacionSuministradordeServiciosBasicos2022.xlsx", "5. Usuarios", 2022),
        (DOWNLOADS / "MemoriaTarifasdeOperaciónSuministradordeServiciosBasicos2023.xlsx", "5. Usuarios", 2023),
        (DOWNLOADS / "MemoriaTarifasdeOperacionSuministradordeServiciosBasicos2024.xlsx", "2. Usuarios", 2024),
        (DOWNLOADS / "MemoriaTarifasdeOperaciónSuministradordeServiciosBasicos2025.xlsx", "2. Usuarios", 2025),
        (DOWNLOADS / "MemoriaTarifasdeOperacionSuministradoradeServiciosBasicos2026 (4).xlsx", "2. Usuarias", 2026),
    ]
    for path, sheet, year in annual_files:
        if path.exists():
            merge_users(data, parse_users_matrix(path, sheet, year))
        else:
            print(f"  MISSING: {path}")

    return data


# ─── PART 2: ENERGY ───────────────────────────────────────────────────────────

def download_finales(year):
    """Download TarifasFinalesdeSuministroBasico{year}.xlsx if not cached."""
    fname = FINALES_CACHE / f"TarifasFinalesdeSuministroBasico{year}.xlsx"
    if fname.exists() and fname.stat().st_size > 100_000:
        return fname
    # Also check Downloads directly
    direct = DOWNLOADS / f"TarifasFinalesdeSuministroBasico{year}.xlsx"
    if direct.exists() and direct.stat().st_size > 100_000:
        return direct
    url = FINALES_BASE.format(year=year)
    print(f"  Downloading {year} from CNE...")
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    try:
        with urllib.request.urlopen(url, context=ctx, timeout=60) as resp:
            data = resp.read()
        if len(data) < 100_000:
            print(f"    WARNING: suspiciously small ({len(data)} bytes), skipping")
            return None
        fname.write_bytes(data)
        print(f"    → {len(data):,} bytes saved")
        return fname
    except Exception as e:
        print(f"    FAILED: {e}")
        return None


def parse_mercado_sheet(ws, year, month_num):
    """
    Extract energy (MWh) by tarifa×division from a Mercado_MES_YYYY sheet.
    Row 4/5: header with Tarifa | Concepto | División | Unidades | value_col
    Returns {div: {tarifa: MWh_value}}
    """
    all_rows = list(ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True))

    # Find header row: has 'Tarifa' (or 'Categoría tarifaria') and 'División' and 'Unidades'
    hdr_idx = None
    for i, row in enumerate(all_rows):
        has_tarifa = any(v in ('Tarifa', 'Categoría tarifaria', 'Categoria tarifaria') for v in row if v)
        has_div = 'División' in row
        has_units = 'Unidades' in row
        if has_tarifa and has_div and has_units:
            hdr_idx = i
            break
    if hdr_idx is None:
        return {}

    hdr = all_rows[hdr_idx]
    # Column labelled 'Tarifa' in 2019-2023, 'Categoría tarifaria' in 2024+
    col_tarifa = next((i for i, v in enumerate(hdr) if v in ('Tarifa', 'Categoría tarifaria', 'Categoria tarifaria')), None)
    col_concepto = next((i for i, v in enumerate(hdr) if v == 'Concepto'), None)
    col_div = next((i for i, v in enumerate(hdr) if v == 'División'), None)
    col_units = next((i for i, v in enumerate(hdr) if v == 'Unidades'), None)
    # Value column is the one after 'Unidades' that has a datetime or is just after
    col_val = col_units + 1 if col_units is not None else None

    if any(c is None for c in [col_tarifa, col_concepto, col_div, col_val]):
        return {}

    result = {}
    for row in all_rows[hdr_idx + 1:]:
        tarifa = row[col_tarifa]
        concepto = row[col_concepto]
        div_raw = row[col_div]
        units = row[col_units] if col_units is not None else None
        val = row[col_val]

        if tarifa not in TARIFA_CODES:
            continue
        if concepto != 'Energía':
            continue
        if units != 'MWh':
            continue
        div = norm_div(div_raw)
        if div is None:
            continue
        if val is None or not isinstance(val, (int, float)):
            continue
        # Accumulate if multiple rows for same div×tarifa
        if div not in result:
            result[div] = {}
        result[div][tarifa] = result[div].get(tarifa, 0) + float(val)

    return result


def parse_energy_year(path, year):
    """
    Parse all Mercado sheets in a Finales file, sum MWh over available months,
    return annual total per div×tarifa.
    """
    print(f"  Parsing energy for {year}: {path.name}")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets = wb.sheetnames

    MONTH_MAP = {
        "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
        "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
    }

    annual = {}
    months_found = []

    for sname in sheets:
        if not sname.startswith("Mercado_"):
            continue
        # Parse month from sheet name like Mercado_ENERO_2019 or Mercado_MARZO_ABRIL_MAYO_2025
        parts = sname.replace("Mercado_", "").split("_")
        # Parts may be multiple months + year at end
        month_parts = [p for p in parts if p in MONTH_MAP]
        if not month_parts:
            continue
        ws = wb[sname]
        month_data = parse_mercado_sheet(ws, year, month_parts[0])
        months_found.extend(month_parts)
        for div, tarifas in month_data.items():
            for tarifa, mwh in tarifas.items():
                annual.setdefault(div, {}).setdefault(tarifa, 0)
                annual[div][tarifa] += mwh

    wb.close()
    n_months = len(set(months_found))
    print(f"    → {len(annual)} divisions, {n_months} month(s) summed: {sorted(set(months_found))}")
    return annual, n_months


def build_energy_ts():
    """Main energy extraction."""
    data = {}
    status = {}

    years_to_try = [2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
    # 2023 and 2026 are already in Downloads
    special = {
        2023: DOWNLOADS / "TarifasFinalesdeSuministroBasico2023.xlsx",
        2026: DOWNLOADS / "TarifasFinalesdeSuministroBasico2026.xlsx",
    }

    for year in years_to_try:
        path = special.get(year) or download_finales(year)
        if path is None or not path.exists():
            print(f"  SKIPPED {year}: file not available")
            status[year] = "not_available"
            continue
        try:
            annual, n_months = parse_energy_year(path, year)
            status[year] = f"ok_{n_months}_months"
            for div, tarifas in annual.items():
                for tarifa, mwh in tarifas.items():
                    set_nested(data, div, tarifa, year, mwh)
        except Exception as e:
            print(f"  ERROR parsing {year}: {e}")
            status[year] = f"error: {e}"

    return data, status


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("PART 1: Extracting USERS")
    print("=" * 60)
    users_data = build_users_ts()

    # Compute sanity totals
    all_user_years = sorted(set(
        y for d in users_data.values()
        for t in d.values()
        for y in t.keys()
    ))
    print(f"\nUser data covers years: {all_user_years}")

    for check_year in ["2022", "2024", "2026"]:
        total = sum(
            v for d in users_data.values()
            for t in d.values()
            for y, v in t.items()
            if y == check_year and v is not None
        )
        print(f"  National total users {check_year}: {total:,.0f}")

    # Collect unique tarifa codes found
    user_tarifas = sorted(set(t for d in users_data.values() for t in d.keys()))
    user_divs = sorted(users_data.keys())
    print(f"\nDivisions in users data ({len(user_divs)}): {user_divs}")
    print(f"Tarifa codes in users data: {user_tarifas}")

    users_json = {
        "meta": {
            "source": SOURCE_URL,
            "description": (
                "Estimated number of users (usuarias/usuarios) of CFE Suministro Básico "
                "by division tarifaria and tarifa category. "
                "2016-2018: December values from monthly projection series. "
                "2022-2026: annual averages ('promedio') as reported. "
                "2019-2021: not available from Operación workbooks (gap). "
                "For overlapping years, the newer file's figure is used."
            ),
            "units": "users (absolute count)",
            "annualization": (
                "2016-2018: December month value from monthly series; "
                "2022-2026: annual average as stated in each year's workbook"
            ),
            "tarifa_codes": TARIFA_CODES,
            "canonical_divisions": CANONICAL_DIVISIONS,
        },
        "years": all_user_years,
        "byDivision": users_data,
    }

    out_users = OUTPUT_DIR / "cfe_users_ts.json"
    with open(out_users, "w", encoding="utf-8") as f:
        json.dump(users_json, f, ensure_ascii=False, separators=(",", ":"))
    print(f"\n✓ Written: {out_users} ({out_users.stat().st_size:,} bytes)")

    print("\n" + "=" * 60)
    print("PART 2: Extracting ENERGY")
    print("=" * 60)
    energy_data, energy_status = build_energy_ts()

    all_energy_years = sorted(set(
        int(y) for d in energy_data.values()
        for t in d.values()
        for y in t.keys()
    ))
    energy_tarifas = sorted(set(t for d in energy_data.values() for t in d.keys()))
    energy_divs = sorted(energy_data.keys())

    print(f"\nEnergy data covers years: {all_energy_years}")
    print(f"Divisions in energy data ({len(energy_divs)}): {energy_divs}")
    print(f"Tarifa codes in energy data: {energy_tarifas}")
    print(f"Download/parse status: {energy_status}")

    # Sanity: national total 2023
    for check_year in ["2023", "2024"]:
        total = sum(
            v for d in energy_data.values()
            for t in d.values()
            for y, v in t.items()
            if y == check_year and v is not None
        )
        print(f"  National total energy {check_year}: {total:,.0f} MWh")

    energy_json = {
        "meta": {
            "source": SOURCE_URL,
            "download_url_pattern": FINALES_BASE,
            "description": (
                "Annual energy sales (ventas de energía) in MWh for CFE Suministro Básico, "
                "by division tarifaria and tarifa category. "
                "Summed across all available monthly 'Mercado_MES_YYYY' sheets. "
                "2021 only has Jan-May; marked in status. "
                "2026 only has Jan-Apr."
            ),
            "units": "MWh",
            "annualization": "Sum of all available monthly Mercado sheets per year",
            "year_status": energy_status,
            "tarifa_codes": TARIFA_CODES,
            "canonical_divisions": CANONICAL_DIVISIONS,
        },
        "years": [str(y) for y in all_energy_years],
        "byDivision": energy_data,
    }

    out_energy = OUTPUT_DIR / "cfe_energy_ts.json"
    with open(out_energy, "w", encoding="utf-8") as f:
        json.dump(energy_json, f, ensure_ascii=False, separators=(",", ":"))
    print(f"\n✓ Written: {out_energy} ({out_energy.stat().st_size:,} bytes)")

    print("\nDONE.")


if __name__ == "__main__":
    main()
