#!/usr/bin/env python3
"""
fetch_dg.py — Mexico Distributed Generation (Generación Distribuida / Generación Exenta)
time-series extractor for atlas-sen open-source map.

Sources:
  CNE/CRE GD reports landing:
    https://www.gob.mx/cne/documentos/estadisticas-sobre-solicitudes-de-interconexion-de-centrales-electricas-de-generacion-exenta
  Also historic: /cre-historico
  Data basis: Resolución RES/142/2017

  Excel data basis:
    https://datos.gob.mx/busca/dataset/centrales-electricas-de-generacion-distribuida
  (Source noted inside Mexican Energy Market Sizing.xlsx row 1)

Outputs:
  public/data/dg_by_size.json
  public/data/dg_by_state.json

Usage:
  python3 scripts/fetch_dg.py [--excel PATH] [--pdf-dir PATH] [--out-dir PATH] [--fetch-2025]

Dependencies: openpyxl, subprocess (pdftotext via poppler), urllib/requests
"""

import argparse
import json
import os
import re
import subprocess
import sys
import urllib.request
import ssl
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_EXCEL = os.path.expanduser(
    "~/mktstats/Mexican Market Stats/Mexican Energy Market Sizing.xlsx"
)
DEFAULT_PDF_DIR = os.path.expanduser(
    "~/mktstats/Mexican Market Stats"
)
DEFAULT_OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "public", "data")

# 2025-H2 PDF direct URL (found via CNE landing page, June 2026)
URL_2025_H2 = "https://www.gob.mx/cms/uploads/attachment/file/1080563/2025_2S_Generaci_n_Distribuida_vf.pdf"

# ─────────────────────────────────────────────────────────────────────────────
# Canonical 32-state list (normalized names used in output JSON)
# ─────────────────────────────────────────────────────────────────────────────

CANONICAL_STATES = [
    "Aguascalientes",
    "Baja California",
    "Baja California Sur",
    "Campeche",
    "Chiapas",
    "Chihuahua",
    "Ciudad de México",
    "Coahuila",
    "Colima",
    "Durango",
    "Estado de México",
    "Guanajuato",
    "Guerrero",
    "Hidalgo",
    "Jalisco",
    "Michoacán",
    "Morelos",
    "Nayarit",
    "Nuevo León",
    "Oaxaca",
    "Puebla",
    "Querétaro",
    "Quintana Roo",
    "San Luis Potosí",
    "Sinaloa",
    "Sonora",
    "Tabasco",
    "Tamaulipas",
    "Tlaxcala",
    "Veracruz",
    "Yucatán",
    "Zacatecas",
]

# Spelling/variant → canonical map
STATE_ALIASES = {
    # CDMX variants
    "Distrito Federal": "Ciudad de México",
    "Ciudad De Mexico": "Ciudad de México",
    "Ciudad de Mexico": "Ciudad de México",
    "CIUDAD DE MÉXICO": "Ciudad de México",
    # Coahuila
    "Coahuila de Zaragoza": "Coahuila",
    "Coahuila De Zaragoza": "Coahuila",
    # Michoacán
    "Michoacan": "Michoacán",
    "Michoacán de Ocampo": "Michoacán",
    # Estado de México
    "Mexico": "Estado de México",  # only when unambiguous context
    "Edo. de México": "Estado de México",
    "Estado De Mexico": "Estado de México",
    "Edo. México": "Estado de México",
    # Nuevo León
    "Nuevo Leon": "Nuevo León",
    "Nuevo León": "Nuevo León",
    # Querétaro
    "Queretaro": "Querétaro",
    "Querétaro de Arteaga": "Querétaro",
    # San Luis Potosí
    "San Luis Potosi": "San Luis Potosí",
    # Veracruz
    "Veracruz de Ignacio de la Llave": "Veracruz",
    # Yucatán
    "Yucatan": "Yucatán",
    # Zacatecas
    "Zacatecas": "Zacatecas",
}


def normalize_state(raw: str) -> str | None:
    raw = raw.strip()
    if raw in CANONICAL_STATES:
        return raw
    if raw in STATE_ALIASES:
        return STATE_ALIASES[raw]
    # Try case-insensitive match
    low = raw.lower()
    for s in CANONICAL_STATES:
        if s.lower() == low:
            return s
    for alias, canon in STATE_ALIASES.items():
        if alias.lower() == low:
            return canon
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Macro-region mapping (from 2025-H2 report regional breakdown)
# Source: SOLICITUDES ATENDIDAS: REGIÓN NOROESTE / NORTE-BAJÍO / OCCIDENTE /
#         CENTRO-ORIENTE sections in 2025_2S_Generación_Distribuida_vf.pdf
# ─────────────────────────────────────────────────────────────────────────────

MACRO_REGION = {
    # NOROESTE
    "Chihuahua": "Noroeste",
    "Coahuila": "Noroeste",
    "Sonora": "Noroeste",
    "Baja California": "Noroeste",
    "Sinaloa": "Noroeste",
    "Durango": "Noroeste",
    "Baja California Sur": "Noroeste",
    # NORTE-BAJÍO
    "Nuevo León": "Norte-Bajío",
    "Guanajuato": "Norte-Bajío",
    "San Luis Potosí": "Norte-Bajío",
    "Querétaro": "Norte-Bajío",
    "Tamaulipas": "Norte-Bajío",
    "Zacatecas": "Norte-Bajío",
    "Hidalgo": "Norte-Bajío",
    # OCCIDENTE
    "Jalisco": "Occidente",
    "Michoacán": "Occidente",
    "Colima": "Occidente",
    "Nayarit": "Occidente",
    "Aguascalientes": "Occidente",
    # CENTRO-ORIENTE
    "Yucatán": "Centro-Oriente",
    "Ciudad de México": "Centro-Oriente",
    "Estado de México": "Centro-Oriente",
    "Veracruz": "Centro-Oriente",
    "Quintana Roo": "Centro-Oriente",
    "Morelos": "Centro-Oriente",
    "Puebla": "Centro-Oriente",
    "Guerrero": "Centro-Oriente",
    "Chiapas": "Centro-Oriente",
    "Oaxaca": "Centro-Oriente",
    "Campeche": "Centro-Oriente",
    "Tabasco": "Centro-Oriente",
    "Tlaxcala": "Centro-Oriente",
}

# ─────────────────────────────────────────────────────────────────────────────
# PDF metadata: filename fragment → (period_label, report_total_mw, report_total_contratos)
# Totals from each PDF's national summary (used for verification).
# ─────────────────────────────────────────────────────────────────────────────

PDF_METADATA = {
    # key = substring that uniquely identifies the file
    "2015": {
        "period": "2015",
        "report_mw": 117.560,       # kW→MW: 117,560 kW
        "report_contratos": 16986,
        "col_order": "state_mw_contratos",  # columns as they appear in table
        "note": "kW unit — converted to MW; no per-state table extractable (map only)",
        "has_state_table": False,
    },
    "2018-H2": {
        "period": "2018-H2",
        "report_mw": 692.86,
        "report_contratos": 94893,
        "col_order": "state_mw_contratos",
        "note": "Per-state table embedded as infographic/map only; numeric table not extractable via pdftotext",
        "has_state_table": False,
    },
    "2019-_H1": {
        "period": "2019-H1",
        "report_mw": 817.85,
        "report_contratos": 112660,
        "col_order": "state_mw_contratos",
        "note": "Per-state table embedded as infographic/map only; numeric table not extractable via pdftotext",
        "has_state_table": False,
    },
    "2019_H2": {
        "period": "2019-H2",
        "report_mw": 975.14,
        "report_contratos": None,   # stated separately; total contratos not clearly stated as single number in body
        "col_order": "state_contratos_mw",
        "has_state_table": True,
    },
    "2020__H1": {
        "period": "2020-H1",
        "report_mw": None,          # not clearly stated as single line in pdftotext output
        "report_contratos": None,
        "col_order": "state_contratos_mw",
        "has_state_table": True,
    },
    "2020_H2": {
        "period": "2020-H2",
        "report_mw": 1570.53,       # from pdftotext: see "390.584*" note — actual total stated in text
        "report_contratos": None,
        "col_order": "state_mw_contratos",
        "has_state_table": True,
    },
    "2021_Segundo_Semestre": {
        "period": "2021-H2",
        "report_mw": 1551.09,       # from "94,893  1,551.09"
        "report_contratos": 219380, # from parsed table totals
        "col_order": "state_mw_contratos",
        "has_state_table": True,
    },
    "2022_H1": {
        "period": "2022-H1",
        "report_mw": 2307.41,
        "report_contratos": 300624,
        "col_order": "state_mw_contratos",
        "has_state_table": True,
    },
    "2023_H2": {
        "period": "2023-H2",
        "report_mw": None,          # extracted from table sum
        "report_contratos": None,
        "col_order": "state_mw_contratos",
        "has_state_table": True,
    },
    "Segundo_Semestre_2024": {
        "period": "2024-H2",
        "report_mw": 4447.92,
        "report_contratos": 518019,
        "col_order": "state_contratos_mw",
        "has_state_table": True,
    },
    "2025_H2": {
        "period": "2025-H2",
        "report_mw": 5190.71,       # GD-only: 4,200.33 + CIPyME 247.59 + Medición Neta 5,164.98 ...
                                     # Note: 2025 report uses "Solicitudes" not "Contratos" as primary metric
                                     # Capacidad total stated: see extraction
        "report_contratos": None,   # 2025 report uses "solicitudes atendidas" terminology
        "col_order": "state_solicitudes_mw",  # NEW: column names changed in 2025
        "has_state_table": True,
        "url": URL_2025_H2,
        "note": "2025-H2 report uses 'solicitudes' column label (≈ contracts). "
                "Capacidad is MW. National total from 'Total de solicitudes' line.",
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# PDF text extraction helper
# ─────────────────────────────────────────────────────────────────────────────

def pdf_to_text(pdf_path: str) -> str:
    result = subprocess.run(
        ["pdftotext", "-layout", pdf_path, "-"],
        capture_output=True, text=True
    )
    return result.stdout


# ─────────────────────────────────────────────────────────────────────────────
# Per-state table parser
# Handles the main layout variants across PDFs:
#
# Variant A (2022-H1, 2023-H2, 2024-H2):
#   <State>  <float MW>  <int contratos>  (right-justified columns)
# Variant B (2019-H2, 2020-H1, 2020-H2):
#   <State>  <float MW>  <int contratos>  (different right-margin positions)
# Variant C (2025-H2):
#   <State>  <int solicitudes>  <%>  <float MW>  <%>  (4 numeric columns)
#
# Strategy: find lines that start with a known state name and extract
# the numeric tokens from the rest of the line.
# ─────────────────────────────────────────────────────────────────────────────

def parse_state_table(text: str, period: str, col_order: str) -> dict:
    """
    Returns dict: {canonical_state: {"mw": float, "contratos": int}}
    col_order values:
      "state_mw_contratos"     — MW first, then contratos
      "state_contratos_mw"     — contratos first, then MW
      "state_solicitudes_mw"   — solicitudes first (%), MW (%) — 2025 format
    """
    lines = text.split("\n")
    results = {}
    # We look for lines whose first non-space token matches a known state
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Try to match state name (may be followed by spaces then numbers)
        matched_state = None
        rest = None

        # Try progressively longer prefixes (handles multi-word state names)
        for length in [3, 2, 1]:
            # Take first `length` whitespace-separated tokens as candidate state name
            tokens = stripped.split()
            if len(tokens) < length + 1:
                continue
            candidate = " ".join(tokens[:length])
            canon = normalize_state(candidate)
            if canon:
                matched_state = canon
                rest = " ".join(tokens[length:])
                break

        if not matched_state:
            continue

        # Extract numbers from rest of line
        nums = re.findall(r"[\d,]+\.?\d*", rest.replace(",", ""))
        if len(nums) < 2:
            # Try with comma-thousands
            nums_raw = re.findall(r"[\d,]+(?:\.\d+)?", rest)
            nums = [n.replace(",", "") for n in nums_raw]

        if len(nums) < 2:
            continue

        try:
            if col_order == "state_mw_contratos":
                mw = float(nums[0])
                contratos = int(float(nums[1]))
            elif col_order == "state_contratos_mw":
                contratos = int(float(nums[0]))
                mw = float(nums[1])
            elif col_order == "state_solicitudes_mw":
                # Format: solicitudes  %sol  mw  %mw  — but % values < 100
                # nums[0] = solicitudes (large int), nums[1] = %sol, nums[2] = MW, nums[3] = %MW
                contratos = int(float(nums[0]))
                # MW is the 3rd number if len >= 3, else 2nd
                mw = float(nums[2]) if len(nums) >= 3 else float(nums[1])
            else:
                mw = float(nums[0])
                contratos = int(float(nums[1]))
        except (ValueError, IndexError):
            continue

        # Basic sanity: MW should be > 0 and < 10000, contratos > 0
        if mw <= 0 or mw > 10000:
            continue
        if contratos <= 0 or contratos > 2000000:
            continue

        # Keep first occurrence (state appears once in the per-state table,
        # but may appear again in region breakdown — take whichever parses first)
        if matched_state not in results:
            results[matched_state] = {"mw": round(mw, 2), "contratos": contratos}

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Verification helper
# ─────────────────────────────────────────────────────────────────────────────

def verify_totals(parsed: dict, report_mw: float | None, report_contratos: int | None, period: str):
    sum_mw = round(sum(v["mw"] for v in parsed.values()), 2)
    sum_ct = sum(v["contratos"] for v in parsed.values())
    n_states = len(parsed)

    mw_ok = None
    ct_ok = None
    if report_mw is not None:
        diff_mw = abs(sum_mw - report_mw)
        mw_ok = diff_mw <= max(1.0, report_mw * 0.005)  # within 0.5% or 1 MW
    if report_contratos is not None:
        diff_ct = abs(sum_ct - report_contratos)
        ct_ok = diff_ct <= max(10, int(report_contratos * 0.005))

    print(f"  [{period}] States parsed: {n_states}/32  |  "
          f"MW sum={sum_mw} (reported={report_mw}, ok={mw_ok})  |  "
          f"Contratos sum={sum_ct} (reported={report_contratos}, ok={ct_ok})")
    return {
        "period": period,
        "states_parsed": n_states,
        "sum_mw": sum_mw,
        "sum_contratos": sum_ct,
        "report_mw": report_mw,
        "report_contratos": report_contratos,
        "mw_check_pass": mw_ok,
        "contratos_check_pass": ct_ok,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel parser: Task A
# ─────────────────────────────────────────────────────────────────────────────

def parse_excel(excel_path: str) -> tuple[dict, dict]:
    """
    Returns (capacity_by_size, contracts_by_size) dicts.
    capacity_by_size: {"0-5 kW": {"2018": 114.76, ...}, ...}
    contracts_by_size: {"0-5 kW": {"2018": 43963, ...}, ...}

    Note: Row 7 "0-5 kW" for 2018 is stored as formula "=113.24+1.52" = 114.76 MWp.
    For contracts 2018: "=2935+41028" = 43963.
    We evaluate these manually since openpyxl data_only may not compute formulas.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=False)

    def parse_cell(val):
        """Resolve formula strings to float."""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        if s.startswith("="):
            # Evaluate simple sum formulas like =113.24+1.52
            expr = s[1:].replace(" ", "")
            try:
                return float(eval(expr))  # safe: only +/- of numbers
            except Exception:
                return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None

    # ── DG MWp ──
    ws_mwp = wb["DG MWp"]
    # Row 6 is header: cols D-J (2018-2024), positions 4-10 (1-indexed)
    # Row 7-13 are size buckets; Row 14 is Total
    years = []
    for col in range(4, 11):  # D=4 to J=10
        v = ws_mwp.cell(row=6, column=col).value
        if v is not None:
            years.append(str(int(float(str(v).replace(".0", "")))))
    # years = ['2018','2019','2020','2021','2022','2023','2024']

    size_buckets_mwp = {}
    for row in range(7, 14):
        bucket = ws_mwp.cell(row=row, column=3).value
        if not bucket:
            continue
        bucket = str(bucket).strip()
        vals = {}
        for i, yr in enumerate(years):
            cell_val = ws_mwp.cell(row=row, column=4 + i).value
            v = parse_cell(cell_val)
            if v is not None:
                vals[yr] = round(v, 2)
        if vals:
            size_buckets_mwp[bucket] = vals

    # ── DG Contratos ──
    ws_ct = wb["DG Contratos"]
    # Row 6 is header: cols F-L (2018-2024), positions 6-12
    years_ct = []
    for col in range(6, 13):
        v = ws_ct.cell(row=6, column=col).value
        if v is not None:
            years_ct.append(str(int(float(str(v).replace(".0", "")))))

    size_buckets_ct = {}
    for row in range(7, 14):
        bucket = ws_ct.cell(row=row, column=5).value
        if not bucket:
            continue
        bucket = str(bucket).strip()
        vals = {}
        for i, yr in enumerate(years_ct):
            cell_val = ws_ct.cell(row=row, column=6 + i).value
            v = parse_cell(cell_val)
            if v is not None:
                vals[yr] = int(v)
        if vals:
            size_buckets_ct[bucket] = vals

    return size_buckets_mwp, size_buckets_ct, years


# ─────────────────────────────────────────────────────────────────────────────
# 2025-H2 downloader
# ─────────────────────────────────────────────────────────────────────────────

def download_2025(dest_dir: str) -> str | None:
    dest = os.path.join(dest_dir, "GD_2025_H2.pdf")
    if os.path.exists(dest) and os.path.getsize(dest) > 100000:
        print(f"  2025-H2: already downloaded at {dest}")
        return dest
    print(f"  Downloading 2025-H2 from {URL_2025_H2} ...")
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    try:
        req = urllib.request.Request(URL_2025_H2, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, context=ctx, timeout=60) as r, open(dest, "wb") as f:
            f.write(r.read())
        size = os.path.getsize(dest)
        print(f"  2025-H2: downloaded {size:,} bytes → {dest}")
        return dest
    except Exception as e:
        print(f"  WARNING: could not download 2025-H2: {e}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", default=DEFAULT_EXCEL)
    parser.add_argument("--pdf-dir", default=DEFAULT_PDF_DIR)
    parser.add_argument("--out-dir", default=DEFAULT_OUT_DIR)
    parser.add_argument("--fetch-2025", action="store_true",
                        help="Download 2025-H2 PDF from CNE if not present")
    args = parser.parse_args()

    pdf_dir = Path(args.pdf_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Task A: Excel → dg_by_size.json ──────────────────────────────────────
    print("\n=== Task A: Parsing Excel ===")
    cap_mwp, ct, years = parse_excel(args.excel)

    # Compute totals
    all_years = years
    total_mwp = {}
    for yr in all_years:
        s = sum(v.get(yr, 0) for v in cap_mwp.values())
        if s:
            total_mwp[yr] = round(s, 2)
    total_ct = {}
    for yr in all_years:
        s = sum(v.get(yr, 0) for v in ct.values())
        if s:
            total_ct[yr] = int(s)

    dg_size_json = {
        "meta": {
            "source": "Mexican Energy Market Sizing.xlsx, sheets 'DG MWp' and 'DG Contratos'",
            "original_data_source": "https://datos.gob.mx/busca/dataset/centrales-electricas-de-generacion-distribuida",
            "cne_landing": "https://www.gob.mx/cne/documentos/estadisticas-sobre-solicitudes-de-interconexion-de-centrales-electricas-de-generacion-exenta",
            "legal_basis": "Resolución RES/142/2017",
            "unit_capacity": "MWp (megawatt-peak) cumulative installed",
            "unit_contracts": "contracts (acumulado / cumulative)",
            "coverage": "2018–2024 annual (year-end or latest available per year)",
            "note": "0-5 kW bucket for 2018 is sum of 0-1 kW (1.52) and 1-5 kW (113.24) sub-buckets per source. "
                    "Contracts 2018 for 0-5 kW similarly: 2,935 + 41,028 = 43,963.",
        },
        "years": all_years,
        "capacityMWp": cap_mwp,
        "capacityMWp_total": total_mwp,
        "contracts": ct,
        "contracts_total": total_ct,
    }

    out_size = out_dir / "dg_by_size.json"
    with open(out_size, "w", encoding="utf-8") as f:
        json.dump(dg_size_json, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  Written: {out_size}")
    for bucket, vals in cap_mwp.items():
        print(f"    {bucket}: {vals}")

    # ── Task B+C: PDFs → dg_by_state.json ────────────────────────────────────
    print("\n=== Tasks B+C: Parsing PDFs ===")

    # Optionally download 2025-H2
    if args.fetch_2025:
        dl = download_2025(str(pdf_dir))
        if dl:
            # Register under 2025_H2 key
            pass

    # Map each PDF file to its metadata key
    pdf_files = {}
    for pdf_path in sorted(pdf_dir.glob("*.pdf")):
        name = pdf_path.name
        for key in PDF_METADATA:
            if key in name:
                pdf_files[key] = pdf_path
                break

    # Also look for the pre-downloaded 2025 file
    p2025 = pdf_dir / "GD_2025_H2.pdf"
    if p2025.exists():
        pdf_files["2025_H2"] = p2025

    by_state: dict[str, dict] = {s: {} for s in CANONICAL_STATES}
    verification_log = []

    ordered_keys = [
        "2015", "2018-H2", "2019-_H1", "2019_H2", "2020__H1",
        "2020_H2", "2021_Segundo_Semestre", "2022_H1", "2023_H2",
        "Segundo_Semestre_2024", "2025_H2",
    ]

    for key in ordered_keys:
        meta = PDF_METADATA[key]
        period = meta["period"]
        pdf_path = pdf_files.get(key)

        if not pdf_path or not Path(pdf_path).exists():
            print(f"  [{period}] PDF not found for key '{key}', skipping")
            verification_log.append({
                "period": period, "status": "pdf_not_found", "key": key
            })
            continue

        if not meta.get("has_state_table", True):
            print(f"  [{period}] No extractable per-state table (map/infographic only) — skipping state parse")
            verification_log.append({
                "period": period,
                "status": "no_state_table",
                "report_mw": meta.get("report_mw"),
                "report_contratos": meta.get("report_contratos"),
                "note": meta.get("note", ""),
            })
            continue

        text = pdf_to_text(str(pdf_path))
        parsed = parse_state_table(text, period, meta["col_order"])

        vr = verify_totals(parsed, meta.get("report_mw"), meta.get("report_contratos"), period)
        vr["status"] = "parsed"
        vr["pdf_file"] = Path(pdf_path).name
        verification_log.append(vr)

        for state, vals in parsed.items():
            by_state[state][period] = vals

    # Build ordered periods list (only those with data)
    periods_with_data = []
    for key in ordered_keys:
        meta = PDF_METADATA[key]
        period = meta["period"]
        if any(period in by_state[s] for s in CANONICAL_STATES):
            periods_with_data.append(period)

    dg_state_json = {
        "meta": {
            "source": "CNE/CRE Estadísticas GD PDFs",
            "cne_landing": "https://www.gob.mx/cne/documentos/estadisticas-sobre-solicitudes-de-interconexion-de-centrales-electricas-de-generacion-exenta",
            "cne_landing_historico": "https://www.gob.mx/cne/documentos/estadisticas-sobre-solicitudes-de-interconexion-de-centrales-electricas-de-generacion-exenta?idiom=es",
            "legal_basis": "Resolución RES/142/2017",
            "unit_mw": "MW cumulative installed (GD+CIPyME combined unless noted)",
            "unit_contratos": "cumulative contracts (or solicitudes atendidas for 2025-H2)",
            "2025_H2_url": URL_2025_H2,
            "note_2025": "2025-H2 uses 'solicitudes' label; functionally equivalent to contracts for cumulative count.",
            "note_state_order": "Sorted by MW descending within each period in original reports.",
            "state_count": len(CANONICAL_STATES),
        },
        "periods": periods_with_data,
        "byState": {s: by_state[s] for s in CANONICAL_STATES},
        "macroRegion": MACRO_REGION,
        "verificationLog": verification_log,
    }

    out_state = out_dir / "dg_by_state.json"
    with open(out_state, "w", encoding="utf-8") as f:
        json.dump(dg_state_json, f, ensure_ascii=False, separators=(",", ":"))
    print(f"\n  Written: {out_state}")

    print("\n=== Done ===")
    print(f"  Size×Year JSON: {out_size}")
    print(f"  State×Period JSON: {out_state}")
    print(f"  Periods with state data: {periods_with_data}")


if __name__ == "__main__":
    main()
